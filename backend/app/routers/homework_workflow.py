import uuid
import openpyxl
import random
import logging
import anyio
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from sqlalchemy.exc import IntegrityError
from typing import List, Optional

from app.core.security import RoleChecker
from app.core.database import get_db
from app.core.storage import get_s3_client
from app.core.config import settings
from app.core.kcs import misconception_code_to_tag, MISCONCEPTION_TAG_IDS
from app.core.problem_selection import resolve_assigned_problems_async

from app.models import (
    User,
    WeeklyTarget,
    Problem,
    Attempt,
    MisconceptionQuestion,
    ProblemMisconception,
    StudentHomeworkProgress,
    StudentMPSession,
    StudentMPAttempt,
    StudentMisconceptionRecord,
    XlsxUpload,
    WeeklyClassSummaryReport
)
from app.schemas.homework_workflow import (
    HomeworkStatusResponse,
    MPSessionStatusResponse,
    MPQuestionResponse,
    MPSubmitRequest,
    MPSubmitResponse,
    ClassReportResponse,
    StudentSummaryReport,
    HeatmapResponse,
    MisconceptionCount,
    StudentDrilldownResponse,
    StudentMPAttemptDetail,
    StudentPSAttemptDetail
)

router = APIRouter(prefix="/homework", tags=["homework"])
logger = logging.getLogger(__name__)

# Helper to normalize datetime to UTC
def as_utc(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt.astimezone(timezone.utc)

# Helper to determine target active state
def is_target_active(target: WeeklyTarget) -> bool:
    now = datetime.now(timezone.utc)
    starts_at = as_utc(target.starts_at)
    deadline = as_utc(target.deadline)
    
    if starts_at and now < starts_at:
        return False
    if deadline and now >= deadline:
        return False
    return True

# Problem membership now lives in app.core.problem_selection
# (resolve_assigned_problems_async) so KC / manual / random modes stay consistent.

async def _questions_for_queue_tag(db: AsyncSession, tag: str) -> list[MisconceptionQuestion]:
    """Bank questions for a queue tag: exact match first, KC-family fallback second.

    The tag vocabulary is split (P5/D3): `problem_misconceptions` may carry specific
    codes ("VA-01") while the bank may be seeded with family tags ("VA") — strict
    equality used to silently return nothing and auto-advance the session. Mirrors
    the tolerant prefix match the PS side already does (attempts.py).
    """
    res = await db.execute(select(MisconceptionQuestion).where(MisconceptionQuestion.misconception_tag == tag))
    questions = list(res.scalars().all())
    if questions:
        return questions
    family = misconception_code_to_tag(tag)
    if family and family != tag:
        res = await db.execute(select(MisconceptionQuestion).where(MisconceptionQuestion.misconception_tag == family))
        questions = list(res.scalars().all())
    return questions


def _triggered_tags_for_answer(question: MisconceptionQuestion, selected_option: str, is_correct: bool) -> list:
    """Tags credited to one MP answer (P5/D2) — diagnostic metadata, never scoring.

    Option-level triggers when authored; else the question-level tag on a wrong
    answer. "Tidak Tahu" (D) is absence of evidence, not a misconception (§7-Q2):
    it scores incorrect but credits no tags — the text_input carries the signal.
    """
    if selected_option == "D":
        return []
    if is_correct:
        return []
    opt_map = question.option_misconceptions
    idx = ord(selected_option) - ord("A")
    if isinstance(opt_map, list) and 0 <= idx < len(opt_map):
        entry = opt_map[idx]
        return [str(t) for t in entry] if isinstance(entry, list) else []
    return [question.misconception_tag]

# Helper to pre-aggregate reporting statistics
async def upsert_class_summary(db: AsyncSession, user_id: uuid.UUID, weekly_target_id: uuid.UUID):
    # Fetch weekly target
    target_res = await db.execute(select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id))
    target = target_res.scalar_one_or_none()
    if not target:
        return

    # 1. MP Score
    mp_session_res = await db.execute(
        select(StudentMPSession).where(
            and_(StudentMPSession.user_id == user_id, StudentMPSession.weekly_target_id == weekly_target_id)
        )
    )
    mp_session = mp_session_res.scalar_one_or_none()
    
    mp_score = None
    if mp_session and mp_session.completed_at:
        # Check attempts for this weekly target
        attempts_res = await db.execute(
            select(StudentMPAttempt).where(
                and_(StudentMPAttempt.user_id == user_id, StudentMPAttempt.weekly_target_id == weekly_target_id)
            )
        )
        attempts = attempts_res.scalars().all()
        
        # Calculate grade: correct attempts / total unique questions answered in queue
        # Since we proceed regardless, each item in tag_queue has attempts.
        correct_count = sum(1 for a in attempts if a.status == "correct")
        total_questions = len(mp_session.tag_queue)
        if total_questions > 0:
            mp_score = round((correct_count / total_questions) * 100.0, 2)
        else:
            mp_score = 100.0

    # 2. PS Score
    problems_res = await db.execute(select(Problem))
    all_problems = problems_res.scalars().all()
    matched = await resolve_assigned_problems_async(db, target, all_problems)
    matched_keys = [p.key for p in matched]
    
    ps_score = None
    if matched_keys:
        solved_stmt = select(Attempt.task_ref).where(
            and_(
                Attempt.user_id == user_id,
                Attempt.passed == True,
                Attempt.task_ref.in_(matched_keys)
            )
        ).distinct()
        solved_res = await db.execute(solved_stmt)
        solved_keys = {row[0] for row in solved_res.all()}
        logger.info("upsert_class_summary user_id=%s matched_keys=%s solved_keys=%s", user_id, matched_keys, solved_keys)
        ps_score = round((len(solved_keys) / len(matched_keys)) * 100.0, 2)


    # 3. Total Attempts & Misconceptions count on PS
    ps_attempts_stmt = select(func.count(Attempt.id)).where(
        and_(
            Attempt.user_id == user_id,
            Attempt.task_ref.in_(matched_keys)
        )
    )
    ps_attempts_res = await db.execute(ps_attempts_stmt)
    total_attempts_made = ps_attempts_res.scalar() or 0

    triggered_stmt = select(func.count(StudentMisconceptionRecord.id)).where(
        and_(
            StudentMisconceptionRecord.user_id == user_id,
            StudentMisconceptionRecord.problem_id.in_([p.id for p in matched]),
            StudentMisconceptionRecord.triggered == 1
        )
    )
    triggered_res = await db.execute(triggered_stmt)
    misconceptions_triggered_count = triggered_res.scalar() or 0

    # 4. Last Active Timestamp
    last_ps_stmt = select(func.max(Attempt.timestamp)).where(
        and_(Attempt.user_id == user_id, Attempt.task_ref.in_(matched_keys))
    )
    last_ps_res = await db.execute(last_ps_stmt)
    last_ps_active = last_ps_res.scalar()

    last_mp_stmt = select(func.max(StudentMPAttempt.timestamp)).where(
        and_(StudentMPAttempt.user_id == user_id, StudentMPAttempt.weekly_target_id == weekly_target_id)
    )
    last_mp_res = await db.execute(last_mp_stmt)
    last_mp_active = last_mp_res.scalar()

    last_active_at = None
    if last_ps_active and last_mp_active:
        last_active_at = max(last_ps_active, last_mp_active)
    elif last_ps_active:
        last_active_at = last_ps_active
    elif last_mp_active:
        last_active_at = last_mp_active

    # Upsert summary
    summary_stmt = select(WeeklyClassSummaryReport).where(
        and_(WeeklyClassSummaryReport.weekly_target_id == weekly_target_id, WeeklyClassSummaryReport.user_id == user_id)
    )
    summary_res = await db.execute(summary_stmt)
    summary = summary_res.scalar_one_or_none()

    if not summary:
        summary = WeeklyClassSummaryReport(
            id=uuid.uuid4(),
            weekly_target_id=weekly_target_id,
            user_id=user_id,
            mp_score=mp_score,
            ps_score=ps_score,
            total_attempts_made=total_attempts_made,
            misconceptions_triggered_count=misconceptions_triggered_count,
            last_active_at=last_active_at
        )
        db.add(summary)
    else:
        summary.mp_score = mp_score
        summary.ps_score = ps_score
        summary.total_attempts_made = total_attempts_made
        summary.misconceptions_triggered_count = misconceptions_triggered_count
        summary.last_active_at = last_active_at
        summary.updated_at = func.now()

    await db.commit()


@router.get("/status", response_model=List[HomeworkStatusResponse])
async def list_homework_status(
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["student", "instructor", "researcher", "rater"]))
):
    """Retrieve overall homework packet statuses for all weekly targets."""
    user_id = uuid.UUID(current_user["id"])
    
    # Get all weekly homework targets
    targets_stmt = select(WeeklyTarget).where(WeeklyTarget.kind == "homework").order_by(WeeklyTarget.week)
    targets_res = await db.execute(targets_stmt)
    targets = targets_res.scalars().all()

    # Get student progress records
    progress_stmt = select(StudentHomeworkProgress).where(StudentHomeworkProgress.user_id == user_id)
    progress_res = await db.execute(progress_stmt)
    progress_records = {p.weekly_target_id: p for p in progress_res.scalars().all()}

    response = []
    for t in targets:
        prog = progress_records.get(t.id)
        
        submitted_at = None
        if prog:
            mp_status = prog.mp_status
            ps_status = prog.ps_status
            submitted_at = as_utc(prog.submitted_at)
        else:
            # Determine active status dynamically
            active = is_target_active(t)
            if active:
                mp_status = "green"
                ps_status = "yellow"
            else:
                mp_status = "red"
                ps_status = "red"

        response.append(
            HomeworkStatusResponse(
                weekly_target_id=t.id,
                week=t.week,
                topic_kc_focus=t.topic_kc_focus,
                title=t.title,
                deadline=as_utc(t.deadline),
                mp_status=mp_status,
                ps_status=ps_status,
                submitted_at=submitted_at,
            )
        )
    return response


@router.get("/{weekly_target_id}/mp-session", response_model=MPSessionStatusResponse)
async def get_or_create_mp_session(
    weekly_target_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["student"]))
):
    """Retrieve the current student's MP session for the weekly target, or create one if it doesn't exist."""
    user_id = uuid.UUID(current_user["id"])

    # 1. Check if the target exists
    target_stmt = select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id)
    target_res = await db.execute(target_stmt)
    target = target_res.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Weekly homework target not found")

    # 2. Check if the target is active (Red checks)
    progress_stmt = select(StudentHomeworkProgress).where(
        and_(StudentHomeworkProgress.user_id == user_id, StudentHomeworkProgress.weekly_target_id == weekly_target_id)
    )
    progress_res = await db.execute(progress_stmt)
    progress = progress_res.scalar_one_or_none()

    if not progress:
        if not is_target_active(target):
            raise HTTPException(status_code=403, detail="This homework target is not currently available.")
        # Initialize progress record as in_progress for MP
        progress = StudentHomeworkProgress(
            id=uuid.uuid4(),
            user_id=user_id,
            weekly_target_id=weekly_target_id,
            mp_status="in_progress",
            ps_status="yellow",
            mp_started_at=func.now()
        )
        db.add(progress)
        try:
            await db.commit()
        except IntegrityError:
            await db.rollback()
            progress_res = await db.execute(progress_stmt)
            progress = progress_res.scalar_one_or_none()
    elif progress.mp_status == "red":
        raise HTTPException(status_code=403, detail="This homework target is locked.")
    elif progress.mp_status == "completed":
        return MPSessionStatusResponse(
            active=False,
            completed=True,
            weekly_target_id=weekly_target_id
        )

    # 3. Check for existing session
    session_stmt = select(StudentMPSession).where(
        and_(StudentMPSession.user_id == user_id, StudentMPSession.weekly_target_id == weekly_target_id)
    )
    session_res = await db.execute(session_stmt)
    session = session_res.scalar_one_or_none()

    if not session:
        # Build queue of misconceptions
        # Find matching problems
        problems_res = await db.execute(select(Problem))
        all_problems = problems_res.scalars().all()
        matched = await resolve_assigned_problems_async(db, target, all_problems)
        matched_ids = [p.id for p in matched]

        tag_queue = []
        if matched_ids:
            # Gather mappings ordered by problem sequence/KC overlap
            mapping_stmt = select(ProblemMisconception).where(ProblemMisconception.problem_id.in_(matched_ids)).order_by(ProblemMisconception.created_at)
            mapping_res = await db.execute(mapping_stmt)
            mappings = mapping_res.scalars().all()
            
            # Keep order of problems, appending their linked misconception tags (Duplication Rule)
            for p in matched:
                p_tags = [m.misconception_tag for m in mappings if m.problem_id == p.id]
                tag_queue.extend(p_tags)

        # Progress Tracking: filter out previously answered correctly
        correct_attempts_stmt = select(StudentMPAttempt.misconception_tag).where(
            and_(StudentMPAttempt.user_id == user_id, StudentMPAttempt.status == "correct")
        ).distinct()
        correct_attempts_res = await db.execute(correct_attempts_stmt)
        correct_tags = {row[0] for row in correct_attempts_res.all()}

        filtered_queue = [t for t in tag_queue if t not in correct_tags]

        # Fallback to random if exhausted or none linked
        if not filtered_queue:
            bank_stmt = select(MisconceptionQuestion.misconception_tag).distinct()
            bank_res = await db.execute(bank_stmt)
            bank_tags = [row[0] for row in bank_res.all()]
            
            unanswered_bank = [t for t in bank_tags if t not in correct_tags]
            if not unanswered_bank:
                unanswered_bank = bank_tags  # fallback to all if student cleared everything

            # Pick 3 random
            if unanswered_bank:
                filtered_queue = random.sample(unanswered_bank, min(3, len(unanswered_bank)))
            else:
                filtered_queue = []

        if not filtered_queue:
            # No questions in bank at all, auto-complete
            progress.mp_status = "completed"
            progress.mp_completed_at = func.now()
            progress.ps_status = "green"
            progress.ps_started_at = func.now()
            await db.commit()
            return MPSessionStatusResponse(
                active=False,
                completed=True,
                weekly_target_id=weekly_target_id
            )

        # Get first question (exact tag, else KC-family fallback — P5/D3)
        first_tag = filtered_queue[0]
        questions = await _questions_for_queue_tag(db, first_tag)
        selected_q = random.choice(questions) if questions else None

        session = StudentMPSession(
            id=uuid.uuid4(),
            user_id=user_id,
            weekly_target_id=weekly_target_id,
            tag_queue=filtered_queue,
            current_index=0,
            current_question_id=selected_q.id if selected_q else None,
            max_attempts=3,
            attempts_on_current=0
        )
        db.add(session)
        
        # Ensure progress status is in_progress
        if progress:
            progress.mp_status = "in_progress"
        try:
            await db.commit()
        except IntegrityError:
            await db.rollback()
            session_res = await db.execute(session_stmt)
            session = session_res.scalar_one_or_none()
            if not session:
                raise
    else:
        # Load question if missing but session active
        if session.current_question_id is None and session.current_index < len(session.tag_queue):
            tag = session.tag_queue[session.current_index]
            questions = await _questions_for_queue_tag(db, tag)
            selected_q = random.choice(questions) if questions else None
            if selected_q:
                session.current_question_id = selected_q.id
                await db.commit()

    # Load current question object
    current_q_obj = None
    if session.current_question_id:
        q_obj_res = await db.execute(select(MisconceptionQuestion).where(MisconceptionQuestion.id == session.current_question_id))
        current_q_obj = q_obj_res.scalar_one_or_none()

    q_resp = None
    if current_q_obj:
        q_resp = MPQuestionResponse(
            id=current_q_obj.id,
            misconception_tag=current_q_obj.misconception_tag,
            text_en=current_q_obj.text_en,
            text_id=current_q_obj.text_id,
            code=current_q_obj.code,
            options_en=current_q_obj.options_en,
            options_id=current_q_obj.options_id
        )

    return MPSessionStatusResponse(
        active=True,
        completed=False,
        weekly_target_id=weekly_target_id,
        tag_queue=session.tag_queue,
        current_index=session.current_index,
        total_questions=len(session.tag_queue),
        attempts_on_current=session.attempts_on_current,
        max_attempts=session.max_attempts,
        current_question=q_resp
    )


@router.post("/{weekly_target_id}/mp-submit", response_model=MPSubmitResponse)
async def submit_mp_answer(
    weekly_target_id: uuid.UUID,
    payload: MPSubmitRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["student"]))
):
    """Submit answer to the current MP question. Validate, log attempt, and advance queue."""
    user_id = uuid.UUID(current_user["id"])

    # 1. Fetch active session
    session_stmt = select(StudentMPSession).where(
        and_(
            StudentMPSession.user_id == user_id,
            StudentMPSession.weekly_target_id == weekly_target_id,
            StudentMPSession.completed_at == None
        )
    )
    session_res = await db.execute(session_stmt)
    session = session_res.scalar_one_or_none()
    
    if not session:
        raise HTTPException(status_code=400, detail="No active misconception session found for this week")

    # Verify matching question
    if session.current_question_id != payload.question_id:
        raise HTTPException(status_code=400, detail="Incorrect question reference submitted")

    # 2. Fetch question
    q_stmt = select(MisconceptionQuestion).where(MisconceptionQuestion.id == payload.question_id)
    q_res = await db.execute(q_stmt)
    question = q_res.scalar_one_or_none()
    if not question:
        raise HTTPException(status_code=404, detail="Question details not found")

    # 3. Check correctness
    is_correct = False
    selected_option = payload.selected_option.upper()
    
    if selected_option in ["A", "B", "C"]:
        opt_idx = ord(selected_option) - ord("A")
        is_correct = (opt_idx == question.answer_index)
    elif selected_option == "D":
        is_correct = False  # Option D (Tidak Tahu) is always treated as incorrect

    # 4. Log attempt — including which misconceptions the chosen option reveals
    # (P5/D2; diagnostic metadata only, never a scoring input).
    status_str = "correct" if is_correct else "incorrect"
    attempt_log = StudentMPAttempt(
        id=uuid.uuid4(),
        user_id=user_id,
        weekly_target_id=weekly_target_id,
        misconception_question_id=payload.question_id,
        misconception_tag=question.misconception_tag,
        selected_option=selected_option,
        text_input=payload.text_input,
        status=status_str,
        triggered_tags=_triggered_tags_for_answer(question, selected_option, is_correct)
    )
    db.add(attempt_log)
    
    # 5. Update session states
    session.attempts_on_current += 1
    
    # Proceed to next question logic
    advance = False
    if is_correct:
        advance = True
    elif session.attempts_on_current >= session.max_attempts:
        advance = True  # reached max allowed attempts on this question, forced to proceed
    else:
        # Proceed to next question immediately on incorrect answer (user's preferred rule!)
        # The user's specification "It is supposed to proceed to the next question [when answered incorrectly]"
        # E.g. they only get 1 attempt per question.
        advance = True

    if advance:
        session.current_index += 1
        session.attempts_on_current = 0
        session.current_question_id = None
        
        # Check if completed session
        if session.current_index >= len(session.tag_queue):
            session.completed_at = func.now()
            
            # Update overall progress
            progress_stmt = select(StudentHomeworkProgress).where(
                and_(StudentHomeworkProgress.user_id == user_id, StudentHomeworkProgress.weekly_target_id == weekly_target_id)
            )
            progress_res = await db.execute(progress_stmt)
            progress = progress_res.scalar_one_or_none()
            if progress:
                progress.mp_status = "completed"
                progress.mp_completed_at = func.now()
                progress.ps_status = "green"  # Unlock Problem Solving!
                progress.ps_started_at = func.now()
            else:
                progress = StudentHomeworkProgress(
                    id=uuid.uuid4(),
                    user_id=user_id,
                    weekly_target_id=weekly_target_id,
                    mp_status="completed",
                    mp_completed_at=func.now(),
                    ps_status="green",
                    ps_started_at=func.now()
                )
                db.add(progress)

    await db.commit()

    # Pre-aggregate data for reports asynchronously/synchronously
    await upsert_class_summary(db, user_id, weekly_target_id)

    # 6. Build response status
    # Load next question if session is still active
    next_q_resp = None
    if not session.completed_at and advance:
        # Load next question (exact tag, else KC-family fallback — P5/D3)
        tag = session.tag_queue[session.current_index]
        questions = await _questions_for_queue_tag(db, tag)
        selected_q = random.choice(questions) if questions else None
        if selected_q:
            session.current_question_id = selected_q.id
            await db.commit()
            next_q_resp = MPQuestionResponse(
                id=selected_q.id,
                misconception_tag=selected_q.misconception_tag,
                text_en=selected_q.text_en,
                text_id=selected_q.text_id,
                code=selected_q.code,
                options_en=selected_q.options_en,
                options_id=selected_q.options_id
            )
    elif not session.completed_at and not advance:
        # Keep same question
        next_q_resp = MPQuestionResponse(
            id=question.id,
            misconception_tag=question.misconception_tag,
            text_en=question.text_en,
            text_id=question.text_id,
            code=question.code,
            options_en=question.options_en,
            options_id=question.options_id
        )

    sess_status = MPSessionStatusResponse(
        active=not bool(session.completed_at),
        completed=bool(session.completed_at),
        weekly_target_id=weekly_target_id,
        tag_queue=session.tag_queue,
        current_index=session.current_index,
        total_questions=len(session.tag_queue),
        attempts_on_current=session.attempts_on_current,
        max_attempts=session.max_attempts,
        current_question=next_q_resp
    )

    return MPSubmitResponse(
        correct=is_correct,
        explanation_en=question.explanation_en,
        explanation_id=question.explanation_id,
        session_status=sess_status
    )


@router.post("/{weekly_target_id}/submit")
async def submit_homework_set(
    weekly_target_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["student"]))
):
    """Explicitly finalize a homework/checkpoint set.

    Idempotent: stamps submitted_at / ps_completed_at only on the first submit and
    never moves them afterwards. Once submitted the set is read-only (enforced in
    the attempts POST), so reviewing can no longer shift the completion timestamp.
    """
    user_id = uuid.UUID(current_user["id"])

    target_res = await db.execute(select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id))
    target = target_res.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")

    # Every assigned problem must be solved before the set can be submitted.
    problems = (await db.execute(select(Problem))).scalars().all()
    assigned = await resolve_assigned_problems_async(db, target, problems)
    assigned_keys = [p.key for p in assigned]
    solved_keys: set[str] = set()
    if assigned_keys:
        solved_res = await db.execute(
            select(Attempt.task_ref).where(
                and_(
                    Attempt.user_id == user_id,
                    Attempt.passed == True,  # noqa: E712
                    Attempt.task_ref.in_(assigned_keys),
                )
            ).distinct()
        )
        solved_keys = {row[0] for row in solved_res.all()}
    if not assigned_keys or len(solved_keys) < len(assigned_keys):
        raise HTTPException(status_code=400, detail="Finish every problem in this set before submitting.")

    prog_res = await db.execute(
        select(StudentHomeworkProgress).where(
            and_(
                StudentHomeworkProgress.user_id == user_id,
                StudentHomeworkProgress.weekly_target_id == weekly_target_id,
            )
        )
    )
    progress = prog_res.scalar_one_or_none()
    if progress is None:
        progress = StudentHomeworkProgress(
            id=uuid.uuid4(),
            user_id=user_id,
            weekly_target_id=weekly_target_id,
            mp_status="completed",
            ps_status="completed",
        )
        db.add(progress)

    already_submitted = progress.submitted_at is not None
    if progress.ps_completed_at is None:
        progress.ps_completed_at = func.now()
    if progress.submitted_at is None:
        progress.submitted_at = func.now()
    progress.ps_status = "completed"

    await db.commit()
    await upsert_class_summary(db, user_id, weekly_target_id)

    return {"status": "ok", "already_submitted": already_submitted, "ps_status": "completed"}


# Analytics / Reporting Endpoints

@router.get("/{weekly_target_id}/class-report", response_model=ClassReportResponse)
async def get_class_report(
    weekly_target_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["instructor", "researcher"]))
):
    """Retrieve overview of student performance for the class report by week."""
    # Fetch weekly target
    target_res = await db.execute(select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id))
    target = target_res.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Weekly target not found")

    # Fetch all students
    students_res = await db.execute(select(User).where(User.role == "student"))
    students = students_res.scalars().all()

    # Fetch progress records
    progress_res = await db.execute(
        select(StudentHomeworkProgress).where(StudentHomeworkProgress.weekly_target_id == weekly_target_id)
    )
    progress_map = {p.user_id: p for p in progress_res.scalars().all()}

    # Fetch summary records
    summary_res = await db.execute(
        select(WeeklyClassSummaryReport).where(WeeklyClassSummaryReport.weekly_target_id == weekly_target_id)
    )
    summary_map = {s.user_id: s for s in summary_res.scalars().all()}

    student_reports = []
    for s in students:
        prog = progress_map.get(s.id)
        summ = summary_map.get(s.id)

        mp_status = prog.mp_status if prog else "not_started"
        ps_status = prog.ps_status if prog else "locked"
        
        mp_score = float(summ.mp_score) if summ and summ.mp_score is not None else None
        ps_score = float(summ.ps_score) if summ and summ.ps_score is not None else None
        
        completed = (mp_status == "completed" and ps_status == "completed")
        last_active = summ.last_active_at if summ else None

        student_reports.append(
            StudentSummaryReport(
                user_id=s.id,
                username=s.username,
                mp_status=mp_status,
                mp_score=mp_score,
                ps_status=ps_status,
                ps_score=ps_score,
                completed=completed,
                last_active_at=last_active
            )
        )

    return ClassReportResponse(
        weekly_target_id=weekly_target_id,
        week=target.week,
        topic_kc_focus=target.topic_kc_focus,
        students=student_reports
    )


@router.get("/{weekly_target_id}/heatmap", response_model=HeatmapResponse)
async def get_heatmap_report(
    weekly_target_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["instructor", "researcher"]))
):
    """Retrieve visual identification of common misconceptions (by Knowledge Component / Misconception type)."""
    target_res = await db.execute(select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id))
    target = target_res.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Weekly target not found")

    # Get problems for this target
    problems_res = await db.execute(select(Problem))
    all_problems = problems_res.scalars().all()
    matched = await resolve_assigned_problems_async(db, target, all_problems)
    matched_ids = [p.id for p in matched]

    misconceptions_list = []
    if matched_ids:
        # Mapped misconception tags
        mapping_stmt = select(ProblemMisconception.misconception_tag).where(ProblemMisconception.problem_id.in_(matched_ids)).distinct()
        mapping_res = await db.execute(mapping_stmt)
        tags = [row[0] for row in mapping_res.all()]

        for tag in tags:
            # Total attempts on matched problems
            total_attempts_stmt = select(func.count(Attempt.id)).where(
                Attempt.task_ref.in_([p.key for p in matched])
            )
            total_attempts_res = await db.execute(total_attempts_stmt)
            total_attempts = total_attempts_res.scalar() or 0

            # Triggered records
            triggered_stmt = select(
                func.sum(StudentMisconceptionRecord.triggered),
                func.count(func.distinct(StudentMisconceptionRecord.user_id))
            ).where(
                and_(
                    StudentMisconceptionRecord.misconception_tag == tag,
                    StudentMisconceptionRecord.problem_id.in_(matched_ids),
                    StudentMisconceptionRecord.triggered == 1
                )
            )
            triggered_res = await db.execute(triggered_stmt)
            row = triggered_res.all()
            triggered_count = row[0][0] or 0 if row else 0
            student_count = row[0][1] or 0 if row else 0

            misconceptions_list.append(
                MisconceptionCount(
                    misconception_tag=tag,
                    triggered_count=triggered_count,
                    total_attempts=total_attempts,
                    student_count=student_count
                )
            )

    return HeatmapResponse(
        weekly_target_id=weekly_target_id,
        week=target.week,
        topic_kc_focus=target.topic_kc_focus,
        misconceptions=misconceptions_list
    )


@router.get("/{weekly_target_id}/drill-down/{user_id}", response_model=StudentDrilldownResponse)
async def get_student_drilldown(
    weekly_target_id: uuid.UUID,
    user_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["instructor", "researcher", "rater"]))
):
    """Retrieve detailed view of individual student interactions."""
    # Fetch user details
    user_res = await db.execute(select(User).where(User.id == user_id))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Student not found")

    target_res = await db.execute(select(WeeklyTarget).where(WeeklyTarget.id == weekly_target_id))
    target = target_res.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Weekly target not found")

    # 1. Fetch MP Attempts
    mp_attempts_stmt = select(StudentMPAttempt, MisconceptionQuestion.text_en).join(
        MisconceptionQuestion, StudentMPAttempt.misconception_question_id == MisconceptionQuestion.id
    ).where(
        and_(StudentMPAttempt.user_id == user_id, StudentMPAttempt.weekly_target_id == weekly_target_id)
    ).order_by(StudentMPAttempt.timestamp)
    
    mp_attempts_res = await db.execute(mp_attempts_stmt)
    mp_details = []
    for row in mp_attempts_res.all():
        attempt, q_text = row
        mp_details.append(
            StudentMPAttemptDetail(
                misconception_tag=attempt.misconception_tag,
                question_text=q_text,
                selected_option=attempt.selected_option,
                text_input=attempt.text_input,
                status=attempt.status,
                timestamp=attempt.timestamp
            )
        )

    # 2. Fetch PS Attempts
    problems_res = await db.execute(select(Problem))
    all_problems = problems_res.scalars().all()
    matched = await resolve_assigned_problems_async(db, target, all_problems)
    matched_keys = [p.key for p in matched]

    ps_details = []
    if matched_keys:
        ps_attempts_stmt = select(Attempt).where(
            and_(Attempt.user_id == user_id, Attempt.task_ref.in_(matched_keys))
        ).order_by(Attempt.timestamp)
        ps_attempts_res = await db.execute(ps_attempts_stmt)
        ps_attempts = ps_attempts_res.scalars().all()

        for a in ps_attempts:
            # Fetch problem title
            prob = next((p for p in matched if p.key == a.task_ref), None)
            prob_title = prob.title if prob else a.task_ref
            
            # Fetch code snippet from storage if possible
            code_str = None
            if a.content_ref:
                try:
                    s3 = get_s3_client()
                    obj = s3.get_object(Bucket=settings.MINIO_BUCKET_NAME, Key=a.content_ref)
                    code_str = obj["Body"].read().decode("utf-8")
                except Exception:
                    code_str = "Code unavailable in MinIO"

            # Fetch triggered misconceptions
            m_rec_stmt = select(StudentMisconceptionRecord.misconception_tag).where(
                and_(
                    StudentMisconceptionRecord.attempt_id == a.id,
                    StudentMisconceptionRecord.triggered == 1
                )
            )
            m_rec_res = await db.execute(m_rec_stmt)
            trig_tags = [r[0] for r in m_rec_res.all()]

            ps_details.append(
                StudentPSAttemptDetail(
                    problem_key=a.task_ref,
                    problem_title=prob_title,
                    passed=a.passed or False,
                    submitted_at=a.timestamp,
                    code=code_str,
                    misconceptions_triggered=trig_tags
                )
            )

    return StudentDrilldownResponse(
        user_id=user.id,
        username=user.username,
        weekly_target_id=weekly_target_id,
        week=target.week,
        mp_attempts=mp_details,
        ps_attempts=ps_details
    )


# XLSX Upload Pipeline

@router.post("/upload-xlsx")
async def upload_and_process_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(RoleChecker(["instructor"]))
):
    """
    Upload XLSX spreadsheet containing courses, participants, quiz lists,
    question banks, participant attempts, and response history. 
    Process and reconcile data synchronously.
    """
    user_id = uuid.UUID(current_user["id"])
    
    # 1. Log upload entry
    upload_record = XlsxUpload(
        id=uuid.uuid4(),
        filename=file.filename,
        uploaded_by=user_id,
        status="processing"
    )
    db.add(upload_record)
    await db.commit()

    # Save to MinIO as audit log
    file_bytes = await file.read()
    try:
        def upload_to_s3():
            s3 = get_s3_client()
            s3.put_object(
                Bucket=settings.MINIO_BUCKET_NAME,
                Key=f"xlsx_uploads/{upload_record.id}.xlsx",
                Body=file_bytes,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        await anyio.to_thread.run_sync(upload_to_s3)
    except Exception as s3_err:
        logger.warning("Failed to store XLSX file in MinIO audit log: %s", s3_err)

    # 2. Parse Excel
    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as parse_err:
        upload_record.status = "failed"
        upload_record.error_message = f"Failed to parse Excel file format: {str(parse_err)}"
        await db.commit()
        raise HTTPException(status_code=400, detail=upload_record.error_message)

    # Summary of processed rows
    reconciled = {
        "participants": 0,
        "weekly_targets": 0,
        "misconception_questions": 0,
        "problem_misconceptions": 0
    }
    # Non-fatal authoring problems (e.g. unknown misconception tags) — reported
    # back to the uploader instead of failing the whole file (P5/R1a).
    warnings: list[str] = []

    try:
        # A. Process Participants / Users
        user_sheet = next((wb[s] for s in wb.sheetnames if s.lower() in ["participants", "peserta", "users"]), None)
        if user_sheet:
            for row in user_sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                username = str(row[0]).strip()
                role = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "student"
                full_name = str(row[2]).strip() if len(row) > 2 and row[2] else username
                
                # Reconcile user in database
                existing_res = await db.execute(select(User).where(User.username == username))
                existing_user = existing_res.scalar_one_or_none()
                if not existing_user:
                    # Create basic user
                    import bcrypt
                    hashed = bcrypt.hashpw(username.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                    new_user = User(
                        id=uuid.uuid4(),
                        username=username,
                        hashed_password=hashed,
                        role=role
                    )
                    db.add(new_user)
                    reconciled["participants"] += 1
                else:
                    existing_user.role = role
                    reconciled["participants"] += 1

        # B. Process Weekly Targets
        target_sheet = next((wb[s] for s in wb.sheetnames if s.lower() in ["quiz lists", "quizzes", "targets", "weekly_targets"]), None)
        if target_sheet:
            for row in target_sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                course_ref = str(row[0]).strip()
                week = int(row[1]) if len(row) > 1 and row[1] is not None else 1
                topic = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                target_task = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                source = str(row[4]).strip() if len(row) > 4 and row[4] else "LMS"
                title = str(row[5]).strip() if len(row) > 5 and row[5] else None
                description = str(row[6]).strip() if len(row) > 6 and row[6] else None
                
                # Check for existing target
                exist_t_res = await db.execute(
                    select(WeeklyTarget).where(
                        and_(WeeklyTarget.course_ref == course_ref, WeeklyTarget.week == week)
                    )
                )
                existing_t = exist_t_res.scalar_one_or_none()
                if not existing_t:
                    new_t = WeeklyTarget(
                        id=uuid.uuid4(),
                        course_ref=course_ref,
                        week=week,
                        topic_kc_focus=topic,
                        target_task=target_task,
                        source=source,
                        title=title,
                        description=description,
                        kind="homework"
                    )
                    db.add(new_t)
                else:
                    existing_t.topic_kc_focus = topic
                    existing_t.target_task = target_task
                    existing_t.source = source
                    existing_t.title = title
                    existing_t.description = description
                reconciled["weekly_targets"] += 1

        # C. Process Question Bank
        q_sheet = next((wb[s] for s in wb.sheetnames if s.lower() in ["question bank", "bank", "misconception_questions", "questions"]), None)
        if q_sheet:
            for row in q_sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                tag = str(row[0]).strip()
                text_en = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                text_id = str(row[2]).strip() if len(row) > 2 and row[2] else text_en
                code_snippet = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                # Parse JSON options or comma-separated options
                def parse_options(val):
                    if not val:
                        return []
                    v_str = str(val).strip()
                    if v_str.startswith("[") and v_str.endswith("]"):
                        import json
                        try:
                            return json.loads(v_str)
                        except Exception:
                            pass
                    return [opt.strip() for opt in v_str.split(",") if opt.strip()]

                opts_en = parse_options(row[4]) if len(row) > 4 else []
                opts_id = parse_options(row[5]) if len(row) > 5 else opts_en
                answer_idx = int(row[6]) if len(row) > 6 and row[6] is not None else 0
                exp_en = str(row[7]).strip() if len(row) > 7 and row[7] else None
                exp_id = str(row[8]).strip() if len(row) > 8 and row[8] else exp_en

                # P5/R1a: optional per-option trigger columns (option_a/b/c_misconceptions,
                # cols 10-12) — comma-separated or JSON-array tag lists. Assembled into a
                # list parallel to opts_en; unknown tags are skipped and reported, a
                # length mismatch is padded with []. Left as None when no cell is filled
                # so re-uploading an old-format file never wipes authored triggers.
                option_misc = None
                trigger_cols = [row[i] if len(row) > i else None for i in (9, 10, 11)]
                if any(c is not None and str(c).strip() for c in trigger_cols):
                    option_misc = []
                    for opt_i in range(len(opts_en)):
                        raw_cell = trigger_cols[opt_i] if opt_i < len(trigger_cols) else None
                        if opt_i == answer_idx:
                            # The correct option carries no triggers by definition.
                            option_misc.append([])
                            continue
                        valid_tags = []
                        for tg in parse_options(raw_cell):
                            tg_norm = str(tg).strip().upper()
                            if not tg_norm:
                                continue
                            if tg_norm in MISCONCEPTION_TAG_IDS or misconception_code_to_tag(tg_norm):
                                valid_tags.append(tg_norm)
                            else:
                                warnings.append(
                                    f"question bank '{text_en[:40]}': unknown tag '{tg_norm}' (option {chr(65 + opt_i)}) skipped"
                                )
                        option_misc.append(valid_tags)

                # Reconcile questions based on tag and text match
                exist_q_res = await db.execute(
                    select(MisconceptionQuestion).where(
                        and_(MisconceptionQuestion.misconception_tag == tag, MisconceptionQuestion.text_en == text_en)
                    )
                )
                existing_q = exist_q_res.scalar_one_or_none()
                if not existing_q:
                    new_q = MisconceptionQuestion(
                        id=uuid.uuid4(),
                        misconception_tag=tag,
                        text_en=text_en,
                        text_id=text_id,
                        code=code_snippet,
                        options_en=opts_en,
                        options_id=opts_id,
                        answer_index=answer_idx,
                        explanation_en=exp_en,
                        explanation_id=exp_id,
                        option_misconceptions=option_misc
                    )
                    db.add(new_q)
                else:
                    existing_q.text_id = text_id
                    existing_q.code = code_snippet
                    existing_q.options_en = opts_en
                    existing_q.options_id = opts_id
                    existing_q.answer_index = answer_idx
                    existing_q.explanation_en = exp_en
                    existing_q.explanation_id = exp_id
                    if option_misc is not None:
                        existing_q.option_misconceptions = option_misc
                reconciled["misconception_questions"] += 1

        # D. Process Problem-Misconception mappings
        mapping_sheet = next((wb[s] for s in wb.sheetnames if s.lower() in ["problem_misconceptions", "mappings", "matrix"]), None)
        if mapping_sheet:
            for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                prob_key = str(row[0]).strip()
                misc_tag = str(row[1]).strip()

                # Find problem in DB
                prob_res = await db.execute(select(Problem).where(Problem.key == prob_key))
                problem = prob_res.scalar_one_or_none()
                if problem:
                    # Check for existing mapping
                    exist_map_res = await db.execute(
                        select(ProblemMisconception).where(
                            and_(ProblemMisconception.problem_id == problem.id, ProblemMisconception.misconception_tag == misc_tag)
                        )
                    )
                    existing_map = exist_map_res.scalar_one_or_none()
                    if not existing_map:
                        new_map = ProblemMisconception(
                            id=uuid.uuid4(),
                            problem_id=problem.id,
                            misconception_tag=misc_tag
                        )
                        db.add(new_map)
                        reconciled["problem_misconceptions"] += 1

        # Complete upload record
        upload_record.status = "completed"
        await db.commit()
    except Exception as run_err:
        await db.rollback()
        upload_record.status = "failed"
        upload_record.error_message = f"Error during data processing: {str(run_err)}"
        await db.commit()
        raise HTTPException(status_code=500, detail=upload_record.error_message)

    return {
        "status": "success",
        "details": reconciled,
        "warnings": warnings
    }
