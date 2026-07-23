"""Parsing and upsert logic for LMS (Moodle) XLSX exports.

Feeds the summary-report feature: `parse_workbook` turns the 6-sheet export
into normalized row dicts keyed by their LMS natural keys, `upsert_lms_data`
writes them idempotently (ON CONFLICT DO UPDATE) and links LMS participants
to local accounts by username/email.

Export quirks this module absorbs:
- numeric IDs arrive as floats (18332.0), Question_ID can be the string "Random"
- datetimes are naive WIB wall-clock; epoch (1970-01-01) means "not set"
- fully empty rows appear at the end of sheets
- the attempt header (start/finish/grade) is repeated on every slot row
- a future QUESTION_BANK "Tag" column will carry misconception tags (VA-01, ...);
  parsed when present, null otherwise
"""

from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from typing import Any, BinaryIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy import select, update, func

from app.models.lms import (
    LmsCourse,
    LmsParticipant,
    LmsQuiz,
    LmsQuestion,
    LmsAttempt,
    LmsAttemptAnswer,
    LmsResponseStep,
)

# WIB is UTC+7 year-round (no DST), so a fixed offset is safe and avoids a tzdata dependency.
WIB = timezone(timedelta(hours=7))

REQUIRED_SHEETS = [
    "COURSES",
    "PARTICIPANTS",
    "QUIZ_LIST",
    "QUESTION_BANK",
    "PARTICIPANT_ATTEMPTS",
    "RESPONSE_HISTORY",
]

# Candidate header names for the upcoming misconception tag column.
TAG_HEADERS = ["Tag", "Tags", "Misconception_Tag", "Misconception_Tags"]

# Rows per INSERT statement; keeps parameter counts well under asyncpg's 32767 cap.
CHUNK_SIZE = 500


def _to_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).replace("\x00", "").strip()  # Postgres rejects NUL bytes in text
    return s or None


def _to_dt(v: Any) -> datetime | None:
    """Naive WIB wall-clock -> tz-aware; epoch sentinel -> None."""
    if not isinstance(v, datetime):
        return None
    if v.year <= 1970:
        return None
    if v.tzinfo is None:
        return v.replace(tzinfo=WIB)
    return v


def _parse_tags(v: Any) -> list[str] | None:
    s = _to_str(v)
    if not s:
        return None
    tags = [t.strip().upper() for t in s.replace(";", ",").split(",") if t.strip()]
    return tags or None


@dataclass
class ParsedLms:
    """Row dicts keyed by natural key (dicts dedupe repeated rows, last wins)."""

    courses: dict[int, dict] = field(default_factory=dict)
    participants: dict[tuple, dict] = field(default_factory=dict)
    quizzes: dict[int, dict] = field(default_factory=dict)
    questions: dict[tuple, dict] = field(default_factory=dict)
    attempts: dict[tuple, dict] = field(default_factory=dict)
    attempt_answers: dict[tuple, dict] = field(default_factory=dict)
    response_steps: dict[tuple, dict] = field(default_factory=dict)

    def row_counts(self) -> dict[str, int]:
        return {
            "courses": len(self.courses),
            "participants": len(self.participants),
            "quizzes": len(self.quizzes),
            "questions": len(self.questions),
            "attempts": len(self.attempts),
            "attempt_answers": len(self.attempt_answers),
            "response_steps": len(self.response_steps),
        }


def _sheet_rows(ws):
    """Yield {header: value} dicts, skipping blank rows; tolerant to short rows."""
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    keys = [_to_str(h) for h in header]
    width = len(keys)
    for raw in rows:
        if not raw or not any(v is not None for v in raw):
            continue
        padded = tuple(raw) + (None,) * (width - len(raw))
        yield {k: padded[i] for i, k in enumerate(keys) if k}


def _ensure_course(parsed: ParsedLms, course_id: int | None, name: Any = None) -> bool:
    """Seed a course stub from any sheet so FKs never dangle; COURSES enriches it."""
    if course_id is None:
        return False
    if course_id not in parsed.courses:
        parsed.courses[course_id] = {"course_id": course_id, "fullname": _to_str(name)}
    return True


def _ensure_quiz(parsed: ParsedLms, quiz_id: int | None, course_id: int | None, name: Any = None) -> bool:
    if quiz_id is None or not _ensure_course(parsed, course_id):
        return False
    if quiz_id not in parsed.quizzes:
        parsed.quizzes[quiz_id] = {"quiz_id": quiz_id, "course_id": course_id, "name": _to_str(name)}
    return True


def _ensure_attempt(parsed: ParsedLms, quiz_id: int, user_id: int, attempt_no: int) -> dict:
    key = (quiz_id, user_id, attempt_no)
    if key not in parsed.attempts:
        parsed.attempts[key] = {
            "quiz_id": quiz_id,
            "lms_user_id": user_id,
            "attempt_number": attempt_no,
            "start_time": None,
            "finish_time": None,
            "total_grade": None,
        }
    return parsed.attempts[key]


def parse_workbook(file: str | BinaryIO) -> ParsedLms:
    """Parse the 6-sheet Moodle export. Raises ValueError on a missing sheet."""
    wb = load_workbook(file, read_only=True, data_only=True)
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Workbook is missing required sheet(s): {', '.join(missing)}")

    parsed = ParsedLms()

    for row in _sheet_rows(wb["COURSES"]):
        course_id = _to_int(row.get("Course_ID"))
        if course_id is None:
            continue
        parsed.courses[course_id] = {
            "course_id": course_id,
            "shortname": _to_str(row.get("Shortname")),
            "fullname": _to_str(row.get("Fullname")),
            "fakultas": _to_str(row.get("Fakultas")),
            "prodi": _to_str(row.get("Prodi")),
            "tahun_ajar": _to_str(row.get("Tahun_Ajar")),
        }

    for row in _sheet_rows(wb["PARTICIPANTS"]):
        course_id = _to_int(row.get("Course_ID"))
        user_id = _to_int(row.get("User_ID"))
        role = _to_str(row.get("Role_Shortname"))
        if user_id is None or role is None or not _ensure_course(parsed, course_id, row.get("Course_Name")):
            continue
        parsed.participants[(course_id, user_id, role)] = {
            "course_id": course_id,
            "lms_user_id": user_id,
            "role_shortname": role,
            "username": _to_str(row.get("Username")),
            "firstname": _to_str(row.get("Firstname")),
            "lastname": _to_str(row.get("Lastname")),
            "email": _to_str(row.get("Email")),
            "role_name": _to_str(row.get("Role_Name")),
        }

    for row in _sheet_rows(wb["QUIZ_LIST"]):
        quiz_id = _to_int(row.get("Quiz_ID"))
        course_id = _to_int(row.get("Course_ID"))
        if quiz_id is None or not _ensure_course(parsed, course_id, row.get("Course_Name")):
            continue
        parsed.quizzes[quiz_id] = {
            "quiz_id": quiz_id,
            "course_id": course_id,
            "name": _to_str(row.get("Quiz_Name")),
            "open_time": _to_dt(row.get("Open_Time")),
            "close_time": _to_dt(row.get("Close_Time")),
            "time_limit_seconds": _to_int(row.get("Time_Limit_Seconds")),
            "max_grade": _to_float(row.get("Max_Grade")),
            "total_questions": _to_int(row.get("Total_Questions")),
        }

    for row in _sheet_rows(wb["QUESTION_BANK"]):
        quiz_id = _to_int(row.get("Quiz_ID"))
        slot = _to_int(row.get("Slot_Number"))
        if slot is None or not _ensure_quiz(parsed, quiz_id, _to_int(row.get("Course_ID")), row.get("Quiz_Name")):
            continue
        tags = None
        for h in TAG_HEADERS:
            if h in row:
                tags = _parse_tags(row.get(h))
                break
        parsed.questions[(quiz_id, slot)] = {
            "quiz_id": quiz_id,
            "slot_number": slot,
            "question_id": _to_int(row.get("Question_ID")),  # None for "Random" placeholders
            "name": _to_str(row.get("Question_Name")),
            "type": _to_str(row.get("Question_Type")),
            "text": _to_str(row.get("Question_Text")),
            "misconception_tags": tags,
        }

    for row in _sheet_rows(wb["PARTICIPANT_ATTEMPTS"]):
        quiz_id = _to_int(row.get("Quiz_ID"))
        user_id = _to_int(row.get("User_ID"))
        attempt_no = _to_int(row.get("Attempt_Number"))
        slot = _to_int(row.get("Slot_Number"))
        if None in (user_id, attempt_no, slot):
            continue
        if not _ensure_quiz(parsed, quiz_id, _to_int(row.get("Course_ID")), row.get("Quiz_Name")):
            continue
        attempt = _ensure_attempt(parsed, quiz_id, user_id, attempt_no)
        for col, value in (
            ("start_time", _to_dt(row.get("Attempt_Start_Time"))),
            ("finish_time", _to_dt(row.get("Attempt_Finish_Time"))),
            ("total_grade", _to_float(row.get("Attempt_Total_Grade"))),
        ):
            if value is not None:
                attempt[col] = value
        parsed.attempt_answers[(quiz_id, user_id, attempt_no, slot)] = {
            "quiz_id": quiz_id,
            "lms_user_id": user_id,
            "attempt_number": attempt_no,
            "slot_number": slot,
            "question_summary": _to_str(row.get("Question_Summary")),
            "right_answer": _to_str(row.get("Right_Answer")),
            "student_answer": _to_str(row.get("Student_Answer")),
            "question_state": _to_str(row.get("Question_State")),
            "question_grade": _to_float(row.get("Question_Grade")),
        }

    for row in _sheet_rows(wb["RESPONSE_HISTORY"]):
        quiz_id = _to_int(row.get("Quiz_ID"))
        user_id = _to_int(row.get("User_ID"))
        attempt_no = _to_int(row.get("Attempt_Number"))
        slot = _to_int(row.get("Slot_Number"))
        step = _to_int(row.get("Step"))
        if None in (user_id, attempt_no, slot, step):
            continue
        if not _ensure_quiz(parsed, quiz_id, _to_int(row.get("Course_ID")), row.get("Quiz_Name")):
            continue
        _ensure_attempt(parsed, quiz_id, user_id, attempt_no)
        parsed.response_steps[(quiz_id, user_id, attempt_no, slot, step)] = {
            "quiz_id": quiz_id,
            "lms_user_id": user_id,
            "attempt_number": attempt_no,
            "slot_number": slot,
            "step": step,
            "question_id": _to_int(row.get("Question_ID")),
            "time": _to_dt(row.get("Time")),
            "action": _to_str(row.get("Action")),
            "state": _to_str(row.get("State")),
            "marks": _to_float(row.get("Marks")),
            "coderunner_output": _to_str(row.get("CodeRunner_Output")),
        }

    wb.close()
    return parsed


async def _upsert(db: AsyncSession, model, rows: list[dict], conflict_cols: list[str]) -> None:
    if not rows:
        return
    for i in range(0, len(rows), CHUNK_SIZE):
        chunk = rows[i : i + CHUNK_SIZE]
        stmt = pg_insert(model).values(chunk)
        update_cols = {
            c.name: stmt.excluded[c.name]
            for c in model.__table__.columns
            if c.name not in conflict_cols and c.name != "id"
        }
        stmt = stmt.on_conflict_do_update(index_elements=conflict_cols, set_=update_cols)
        await db.execute(stmt)


async def _match_participants(db: AsyncSession) -> int:
    """Link LMS participants to local accounts by users.username == LMS username or email.

    Returns the number of *student* participants still unmatched.
    """
    from app.models.user import User

    await db.execute(
        update(LmsParticipant)
        .values(matched_user_id=select(User.id)
                .where((User.username == LmsParticipant.username) | (User.username == LmsParticipant.email))
                .limit(1)
                .scalar_subquery())
        .where(LmsParticipant.matched_user_id.is_(None))
    )
    result = await db.execute(
        select(func.count()).select_from(LmsParticipant).where(
            LmsParticipant.role_shortname == "student",
            LmsParticipant.matched_user_id.is_(None),
        )
    )
    return result.scalar_one()


async def upsert_lms_data(db: AsyncSession, parsed: ParsedLms) -> dict:
    """Idempotently write a parsed export. Commits once; returns counts + unmatched."""
    await _upsert(db, LmsCourse, list(parsed.courses.values()), ["course_id"])
    await _upsert(db, LmsParticipant, list(parsed.participants.values()),
                  ["course_id", "lms_user_id", "role_shortname"])
    await _upsert(db, LmsQuiz, list(parsed.quizzes.values()), ["quiz_id"])
    await _upsert(db, LmsQuestion, list(parsed.questions.values()), ["quiz_id", "slot_number"])
    await _upsert(db, LmsAttempt, list(parsed.attempts.values()),
                  ["quiz_id", "lms_user_id", "attempt_number"])
    await _upsert(db, LmsAttemptAnswer, list(parsed.attempt_answers.values()),
                  ["quiz_id", "lms_user_id", "attempt_number", "slot_number"])
    await _upsert(db, LmsResponseStep, list(parsed.response_steps.values()),
                  ["quiz_id", "lms_user_id", "attempt_number", "slot_number", "step"])

    unmatched = await _match_participants(db)
    await db.commit()

    return {"row_counts": parsed.row_counts(), "unmatched_students": unmatched}
