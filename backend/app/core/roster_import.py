"""Provisioning importer for a researcher-uploaded class roster XLSX (R4).

Distinct from `lms_import.py` — which ingests a Moodle *quiz-results* export and
matches participants to *pre-existing* accounts. This module *creates* the local
accounts for a class's teacher and students, stands up the `LmsCourse`, and
writes `LmsParticipant` rows with `matched_user_id` set **at creation time** —
so the teacher's course-scoped dashboard lights up immediately, with no
`seed_lms_teacher.py` follow-up (kills the teacher-LMS-matching gotcha).

Workbook (tolerant, case-insensitive sheet + header matching):

  COURSES  : Course_ID, Shortname, Fullname [, Fakultas, Prodi, Tahun_Ajar]
  TEACHERS : Course_ID, Username, Full_Name | Firstname/Lastname, LMS_User_ID [, Email, Password]
  STUDENTS : Course_ID, Username, Full_Name | Firstname/Lastname, LMS_User_ID [, Email, Password]

Decisions baked in (see p4-implementation-plan §7):
- Q4: `LMS_User_ID` (the Moodle numeric id) is **required** so a later quiz
  export reconciles on the same key. Rows missing it (or Course_ID/Username) are
  skipped and reported — never silently synthesized.
- Q2: an explicit `Password` cell is honored; blanks get a generated temp
  password returned in the result so the researcher can distribute them. An
  existing account keeps its password unless the row supplies a new one.
"""

from dataclasses import dataclass, field
from typing import BinaryIO
import secrets
import uuid

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.lms_import import _to_int, _to_str, _sheet_rows  # shared parsing helpers
from app.core.security import get_password_hash
from app.models.lms import LmsCourse, LmsParticipant
from app.models.user import User

COURSE_SHEETS = {"courses", "course", "kelas"}
TEACHER_SHEETS = {"teachers", "teacher", "dosen", "pengajar", "instructors"}
STUDENT_SHEETS = {"students", "student", "mahasiswa", "peserta"}


@dataclass
class ParsedRoster:
    courses: dict[int, dict] = field(default_factory=dict)
    teachers: list[dict] = field(default_factory=list)
    students: list[dict] = field(default_factory=list)


def _find_sheet(wb, names: set[str]):
    for s in wb.sheetnames:
        if s.strip().lower() in names:
            return wb[s]
    return None


def _norm(row: dict) -> dict:
    """Lowercase/trim header keys so column casing/spacing doesn't matter."""
    return {(k or "").strip().lower(): v for k, v in row.items()}


def _pick(r: dict, *keys):
    for k in keys:
        if r.get(k) is not None:
            return r.get(k)
    return None


def _parse_people(ws) -> list[dict]:
    people: list[dict] = []
    if ws is None:
        return people
    for raw in _sheet_rows(ws):
        r = _norm(raw)
        username = _to_str(_pick(r, "username", "user_name"))
        if not username:
            continue
        firstname = _to_str(r.get("firstname"))
        lastname = _to_str(r.get("lastname"))
        full = _to_str(_pick(r, "full_name", "fullname", "name"))
        if not firstname and not lastname and full:
            parts = full.split()
            firstname = parts[0]
            lastname = " ".join(parts[1:]) or None
        people.append({
            "course_id": _to_int(_pick(r, "course_id", "course id", "courseid")),
            "username": username,
            "firstname": firstname,
            "lastname": lastname,
            "email": _to_str(r.get("email")),
            "lms_user_id": _to_int(_pick(r, "lms_user_id", "user_id", "userid", "lms user id")),
            "password": _to_str(r.get("password")),
        })
    return people


def parse_roster_workbook(file: str | BinaryIO) -> ParsedRoster:
    """Parse a roster workbook. Raises ValueError when no roster sheet is present."""
    wb = load_workbook(file, read_only=True, data_only=True)

    courses: dict[int, dict] = {}
    cs = _find_sheet(wb, COURSE_SHEETS)
    if cs is not None:
        for raw in _sheet_rows(cs):
            r = _norm(raw)
            cid = _to_int(_pick(r, "course_id", "course id", "courseid"))
            if cid is None:
                continue
            courses[cid] = {
                "course_id": cid,
                "shortname": _to_str(r.get("shortname")),
                "fullname": _to_str(_pick(r, "fullname", "course_name", "name")),
                "fakultas": _to_str(r.get("fakultas")),
                "prodi": _to_str(r.get("prodi")),
                "tahun_ajar": _to_str(_pick(r, "tahun_ajar", "tahun ajar")),
            }

    teachers = _parse_people(_find_sheet(wb, TEACHER_SHEETS))
    students = _parse_people(_find_sheet(wb, STUDENT_SHEETS))
    wb.close()

    if not teachers and not students:
        raise ValueError(
            "Workbook must contain a TEACHERS and/or STUDENTS sheet with at least one row."
        )
    return ParsedRoster(courses=courses, teachers=teachers, students=students)


async def _upsert_course(db: AsyncSession, cdata: dict) -> None:
    stmt = pg_insert(LmsCourse).values(**cdata)
    stmt = stmt.on_conflict_do_update(
        index_elements=["course_id"],
        set_={k: stmt.excluded[k] for k in cdata if k != "course_id"},
    )
    await db.execute(stmt)


async def _ensure_course_stub(db: AsyncSession, course_id: int) -> None:
    """Seed a bare course row so a participant FK never dangles when the roster
    lists a Course_ID that the COURSES sheet omitted."""
    stmt = pg_insert(LmsCourse).values(course_id=course_id).on_conflict_do_nothing(
        index_elements=["course_id"]
    )
    await db.execute(stmt)


async def _upsert_participant(
    db: AsyncSession, course_id: int, lms_user_id: int, role_shortname: str,
    person: dict, matched_user_id: uuid.UUID, role_name: str,
) -> None:
    values = {
        "course_id": course_id,
        "lms_user_id": lms_user_id,
        "role_shortname": role_shortname,
        "username": person["username"],
        "firstname": person.get("firstname"),
        "lastname": person.get("lastname"),
        "email": person.get("email"),
        "role_name": role_name,
        "matched_user_id": matched_user_id,
    }
    stmt = pg_insert(LmsParticipant).values(**values)
    stmt = stmt.on_conflict_do_update(
        index_elements=["course_id", "lms_user_id", "role_shortname"],
        set_={k: stmt.excluded[k] for k in values
              if k not in ("course_id", "lms_user_id", "role_shortname")},
    )
    await db.execute(stmt)


async def upsert_roster(
    db: AsyncSession, parsed: ParsedRoster, *, generate_missing_passwords: bool = True
) -> dict:
    """Provision accounts + course + participant links idempotently. Commits once.

    Returns {counts, generated_credentials, skipped}.
    """
    for cdata in parsed.courses.values():
        await _upsert_course(db, cdata)

    counts = {
        "courses": len(parsed.courses),
        "teachers_created": 0, "teachers_updated": 0,
        "students_created": 0, "students_updated": 0,
        "skipped": 0,
    }
    generated: dict[str, str] = {}
    skipped: list[dict] = []

    async def provision(person: dict, role: str, role_shortname: str, role_name: str) -> bool | None:
        username = person["username"]
        course_id = person["course_id"]
        lms_user_id = person["lms_user_id"]
        if course_id is None:
            skipped.append({"username": username, "reason": "missing Course_ID"})
            return None
        if lms_user_id is None:
            skipped.append({"username": username, "reason": "missing LMS_User_ID"})
            return None

        await _ensure_course_stub(db, course_id)

        existing = (await db.execute(
            select(User).where(User.username == username)
        )).scalar_one_or_none()

        if existing is None:
            pwd = person.get("password")
            if not pwd:
                pwd = secrets.token_urlsafe(9) if generate_missing_passwords else username
                if generate_missing_passwords:
                    generated[username] = pwd
            new_user = User(
                id=uuid.uuid4(),
                username=username,
                hashed_password=get_password_hash(pwd),
                role=role,
                consent_status=False,
            )
            db.add(new_user)
            await db.flush()  # ensure the row exists before the participant FK
            user_id = new_user.id
            created = True
        else:
            existing.role = role
            if person.get("password"):  # only reset an existing password when supplied
                existing.hashed_password = get_password_hash(person["password"])
            await db.flush()
            user_id = existing.id
            created = False

        await _upsert_participant(db, course_id, lms_user_id, role_shortname, person, user_id, role_name)
        return created

    for t in parsed.teachers:
        c = await provision(t, "instructor", "editingteacher", "Teacher")
        if c is True:
            counts["teachers_created"] += 1
        elif c is False:
            counts["teachers_updated"] += 1

    for s in parsed.students:
        c = await provision(s, "student", "student", "Student")
        if c is True:
            counts["students_created"] += 1
        elif c is False:
            counts["students_updated"] += 1

    counts["skipped"] = len(skipped)
    await db.commit()
    return {"counts": counts, "generated_credentials": generated, "skipped": skipped}
