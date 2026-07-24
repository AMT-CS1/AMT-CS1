import pytest
import uuid
import io
import asyncio
from datetime import datetime, timezone, timedelta
import httpx
import openpyxl
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
from sqlalchemy import select, delete, and_

from app.main import app
from app.core.database import get_db, async_session_maker, engine
from app.core.security import create_access_token
from app.models import (
    User,
    WeeklyTarget,
    Problem,
    Attempt,
    MisconceptionQuestion,
    ProblemMisconception,
    StudentHomeworkProgress,
    StudentMPSession,
    StudentMPAttempt,
    StudentMisconceptionRecord,
    WeeklyClassSummaryReport
)

def get_instructor_headers():
    token = create_access_token({
        "sub": "22222222-2222-2222-2222-222222222222",
        "username": "instructor_user",
        "email": "instructor@example.com",
        "role": "instructor",
        "consent_status": True
    })
    return {"Authorization": f"Bearer {token}"}

def get_student_headers(student_id="11111111-1111-1111-1111-111111111111"):
    token = create_access_token({
        "sub": student_id,
        "username": f"student_{student_id[:8]}",
        "email": f"student_{student_id[:8]}@example.com",
        "role": "student",
        "consent_status": True
    })
    return {"Authorization": f"Bearer {token}"}

def test_homework_workflow_endpoints():
    async def run_test():
        async with async_session_maker() as session:
            await session.begin()
            
            async def override_get_db():
                yield session
                
            app.dependency_overrides[get_db] = override_get_db
            
            # Setup IDs
            target_id = uuid.uuid4()
            problem_id_1 = uuid.uuid4()
            problem_id_2 = uuid.uuid4()
            mq_id_1 = uuid.uuid4()
            mq_id_2 = uuid.uuid4()
            student_id = "11111111-1111-1111-1111-111111111111"

            try:
                # 1. Setup seed data in session
                # A. Create a weekly target
                target = WeeklyTarget(
                    id=target_id,
                    course_ref="CS1-TEST-HW",
                    week=1,
                    topic_kc_focus="VA,OP",
                    target_task="Homework 1",
                    source="test",
                    title="Week 1 Homework",
                    starts_at=datetime.utcnow() - timedelta(days=1),
                    deadline=datetime.utcnow() + timedelta(days=2),

                    kind="homework"
                )
                session.add(target)

                # B. Create Problems (PS)
                problem1 = Problem(
                    id=problem_id_1,
                    key="PS1",
                    title="Problem 1",
                    description_en="Swap vars",
                    description_id="Swap vars",
                    starter_code="x <- 5",
                    test_cases=[],
                    kc_tags="VA,OP"
                )
                problem2 = Problem(
                    id=problem_id_2,
                    key="PS2",
                    title="Problem 2",
                    description_en="Add vars",
                    description_id="Add vars",
                    starter_code="y <- 10",
                    test_cases=[],
                    kc_tags="VA"
                )
                session.add(problem1)
                session.add(problem2)

                # C. Map Problems to Misconceptions (PS1 -> VA-01, OP-01; PS2 -> VA-01)
                pm1 = ProblemMisconception(id=uuid.uuid4(), problem_id=problem_id_1, misconception_tag="VA-01")
                pm2 = ProblemMisconception(id=uuid.uuid4(), problem_id=problem_id_1, misconception_tag="OP-01")
                pm3 = ProblemMisconception(id=uuid.uuid4(), problem_id=problem_id_2, misconception_tag="VA-01")
                session.add(pm1)
                session.add(pm2)
                session.add(pm3)

                # D. Create Misconception Questions (MP Bank)
                mq1 = MisconceptionQuestion(
                    id=mq_id_1,
                    misconception_tag="VA-01",
                    text_en="What is VA-01?",
                    text_id="Apa itu VA-01?",
                    options_en=["Wrong1", "Right1", "Wrong2"],
                    options_id=["Salah1", "Benar1", "Salah2"],
                    answer_index=1,  # B
                    explanation_en="Explanation 1",
                    explanation_id="Penjelasan 1"
                )
                mq2 = MisconceptionQuestion(
                    id=mq_id_2,
                    misconception_tag="OP-01",
                    text_en="What is OP-01?",
                    text_id="Apa itu OP-01?",
                    options_en=["Right2", "Wrong1", "Wrong2"],
                    options_id=["Benar2", "Salah1", "Salah2"],
                    answer_index=0,  # A
                    explanation_en="Explanation 2",
                    explanation_id="Penjelasan 2"
                )
                session.add(mq1)
                session.add(mq2)

                # Create user in case they don't exist
                user_stmt = select(User).where(User.id == uuid.UUID(student_id))
                user_res = await session.execute(user_stmt)
                user = user_res.scalar_one_or_none()
                if not user:
                    user = User(
                        id=uuid.UUID(student_id),
                        username="student_hw_user",
                        hashed_password="password",
                        role="student"
                    )

                    session.add(user)

                await session.commit()

                transport = httpx.ASGITransport(app=app)
                async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
                    # 2. Test status endpoint
                    status_res = await client.get("/homework/status", headers=get_student_headers())
                    assert status_res.status_code == 200
                    status_data = status_res.json()
                    assert len(status_data) >= 1
                    target_status = [t for t in status_data if t["weekly_target_id"] == str(target_id)][0]
                    assert target_status["mp_status"] == "green"
                    assert target_status["ps_status"] == "yellow"

                    # 3. Test GET or create MP session
                    session_res = await client.get(f"/homework/{target_id}/mp-session", headers=get_student_headers())
                    assert session_res.status_code == 200
                    sess_data = session_res.json()
                    assert sess_data["active"] is True
                    assert sess_data["completed"] is False
                    # Duplication rule: VA-01 (PS1), OP-01 (PS1), VA-01 (PS2) => length 3
                    assert len(sess_data["tag_queue"]) == 3
                    assert sess_data["tag_queue"] == ["VA-01", "OP-01", "VA-01"]
                    assert sess_data["current_index"] == 0
                    assert sess_data["current_question"]["misconception_tag"] == "VA-01"

                    # 4. Submit incorrect option D (Tidak Tahu) on question 1
                    sub_d_res = await client.post(
                        f"/homework/{target_id}/mp-submit",
                        json={
                            "weekly_target_id": str(target_id),
                            "question_id": sess_data["current_question"]["id"],
                            "selected_option": "D",
                            "text_input": "I have no idea about variables"
                        },
                        headers=get_student_headers()
                    )
                    assert sub_d_res.status_code == 200
                    sub_d_data = sub_d_res.json()
                    assert sub_d_data["correct"] is False
                    # Incorrect answer must proceed to the next question immediately
                    assert sub_d_data["session_status"]["current_index"] == 1
                    assert sub_d_data["session_status"]["current_question"]["misconception_tag"] == "OP-01"

                    # 5. Submit correct option A on question 2 (OP-01, answer_index=0)
                    sub_c_res = await client.post(
                        f"/homework/{target_id}/mp-submit",
                        json={
                            "weekly_target_id": str(target_id),
                            "question_id": sub_d_data["session_status"]["current_question"]["id"],
                            "selected_option": "A"
                        },
                        headers=get_student_headers()
                    )
                    assert sub_c_res.status_code == 200
                    sub_c_data = sub_c_res.json()
                    assert sub_c_data["correct"] is True
                    assert sub_c_data["session_status"]["current_index"] == 2
                    assert sub_c_data["session_status"]["current_question"]["misconception_tag"] == "VA-01"

                    # 6. Submit correct option B on question 3 (VA-01, answer_index=1) to finish
                    sub_f_res = await client.post(
                        f"/homework/{target_id}/mp-submit",
                        json={
                            "weekly_target_id": str(target_id),
                            "question_id": sub_c_data["session_status"]["current_question"]["id"],
                            "selected_option": "B"
                        },
                        headers=get_student_headers()
                    )
                    assert sub_f_res.status_code == 200
                    sub_f_data = sub_f_res.json()
                    assert sub_f_data["correct"] is True
                    assert sub_f_data["session_status"]["completed"] is True
                    assert sub_f_data["session_status"]["active"] is False

                    # 7. Check status endpoint again (MP should be completed, PS should be green/unlocked)
                    status_res_2 = await client.get("/homework/status", headers=get_student_headers())
                    assert status_res_2.status_code == 200
                    status_data_2 = status_res_2.json()
                    target_status_2 = [t for t in status_data_2 if t["weekly_target_id"] == str(target_id)][0]
                    assert target_status_2["mp_status"] == "completed"
                    assert target_status_2["ps_status"] == "green"

                    # 8. Test Class Report (Instructor role)
                    report_res = await client.get(f"/homework/{target_id}/class-report", headers=get_instructor_headers())
                    assert report_res.status_code == 200
                    report_data = report_res.json()
                    assert report_data["weekly_target_id"] == str(target_id)
                    student_report = [s for s in report_data["students"] if s["user_id"] == student_id][0]
                    assert student_report["mp_status"] == "completed"
                    assert student_report["mp_score"] == pytest.approx(66.67, 0.01) # 2/3 correct

                    # 9. Test Heatmap (Instructor role)
                    heatmap_res = await client.get(f"/homework/{target_id}/heatmap", headers=get_instructor_headers())
                    assert heatmap_res.status_code == 200
                    heatmap_data = heatmap_res.json()
                    assert heatmap_data["weekly_target_id"] == str(target_id)

                    # 10. Test Drilldown
                    drill_res = await client.get(f"/homework/{target_id}/drill-down/{student_id}", headers=get_instructor_headers())
                    assert drill_res.status_code == 200
                    drill_data = drill_res.json()
                    assert len(drill_data["mp_attempts"]) == 3
                    assert drill_data["mp_attempts"][0]["selected_option"] == "D"
                    assert drill_data["mp_attempts"][0]["status"] == "incorrect"

            finally:
                # Cleanup database records
                await session.execute(delete(StudentMisconceptionRecord).where(StudentMisconceptionRecord.user_id == uuid.UUID(student_id)))
                await session.execute(delete(StudentMPAttempt).where(StudentMPAttempt.user_id == uuid.UUID(student_id)))
                await session.execute(delete(StudentMPSession).where(StudentMPSession.user_id == uuid.UUID(student_id)))
                await session.execute(delete(StudentHomeworkProgress).where(StudentHomeworkProgress.user_id == uuid.UUID(student_id)))
                await session.execute(delete(WeeklyClassSummaryReport).where(WeeklyClassSummaryReport.user_id == uuid.UUID(student_id)))
                await session.execute(delete(ProblemMisconception).where(ProblemMisconception.misconception_tag.in_(["VA-01", "OP-01"])))
                await session.execute(delete(MisconceptionQuestion).where(MisconceptionQuestion.id.in_([mq_id_1, mq_id_2])))
                await session.execute(delete(Problem).where(Problem.id.in_([problem_id_1, problem_id_2])))
                await session.execute(delete(WeeklyTarget).where(WeeklyTarget.id == target_id))
                await session.commit()
                
                app.dependency_overrides.pop(get_db, None)
                await engine.dispose()

    asyncio.run(run_test())


def test_xlsx_upload_pipeline():
    async def run_test():
        async with async_session_maker() as session:
            await session.begin()
            
            async def override_get_db():
                yield session
                
            app.dependency_overrides[get_db] = override_get_db
            
            test_username = f"xlsx_user_{uuid.uuid4().hex[:8]}"
            test_week = 999
            
            try:
                # 1. Generate a mock Excel workbook in memory
                wb = openpyxl.Workbook()
                
                # Participants sheet
                ws_u = wb.active
                ws_u.title = "participants"
                ws_u.append(["username", "role", "full_name"])
                ws_u.append([test_username, "student", "XLSX Student Name"])
                
                # Targets sheet
                ws_t = wb.create_sheet(title="weekly_targets")
                ws_t.append(["course_ref", "week", "topic_kc_focus", "target_task", "source", "title", "description"])
                ws_t.append(["CS-XLSX-TEST", test_week, "VA", "XLSX Task", "LMS", "XLSX Week Title", "XLSX Description"])
                
                # Question bank sheet
                ws_q = wb.create_sheet(title="question bank")
                ws_q.append(["misconception_tag", "text_en", "text_id", "code", "options_en", "options_id", "answer_index", "explanation_en", "explanation_id"])
                ws_q.append(["VA-XLSX", "Text EN", "Text ID", "x <- y", '["A", "B", "C"]', '["A", "B", "C"]', 1, "Explanation EN", "Explanation ID"])
                
                # Save to bytes
                out_stream = io.BytesIO()
                wb.save(out_stream)
                out_stream.seek(0)

                # 2. Call upload API
                transport = httpx.ASGITransport(app=app)
                async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
                    files = {"file": ("test.xlsx", out_stream.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                    response = await client.post("/homework/upload-xlsx", files=files, headers=get_instructor_headers())
                    assert response.status_code == 200
                    data = response.json()
                    assert data["status"] == "success"
                    
                    # Verify database entries were populated
                    user_stmt = select(User).where(User.username == test_username)
                    user_res = await session.execute(user_stmt)
                    db_user = user_res.scalar_one_or_none()
                    assert db_user.username == test_username


                    target_stmt = select(WeeklyTarget).where(
                        and_(WeeklyTarget.course_ref == "CS-XLSX-TEST", WeeklyTarget.week == test_week)
                    )
                    target_res = await session.execute(target_stmt)
                    db_target = target_res.scalar_one_or_none()
                    assert db_target is not None
                    assert db_target.topic_kc_focus == "VA"
                    
                    q_stmt = select(MisconceptionQuestion).where(
                        and_(MisconceptionQuestion.misconception_tag == "VA-XLSX", MisconceptionQuestion.text_en == "Text EN")
                    )
                    q_res = await session.execute(q_stmt)
                    db_q = q_res.scalar_one_or_none()
                    assert db_q is not None
                    assert db_q.explanation_en == "Explanation EN"

            finally:
                # Cleanup generated database records
                await session.execute(delete(User).where(User.username == test_username))
                await session.execute(delete(WeeklyTarget).where(WeeklyTarget.course_ref == "CS-XLSX-TEST"))
                await session.execute(delete(MisconceptionQuestion).where(MisconceptionQuestion.misconception_tag == "VA-XLSX"))
                await session.commit()
                
                app.dependency_overrides.pop(get_db, None)
                await engine.dispose()

    asyncio.run(run_test())
