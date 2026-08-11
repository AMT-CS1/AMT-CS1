"""Generate the MP-question authoring template (P5/R1a).

Writes an .xlsx with a `question bank` sheet (the 9 columns the instructor
importer already parses, plus the P5 per-option misconception-trigger columns)
and a `problem_misconceptions` sheet so new items can be wired to PS problems
in the same file. Uploaded through the existing instructor
`POST /homework/upload-xlsx` — no dedicated endpoint.

Run inside the backend container (openpyxl lives there):

    docker exec amt-backend python scripts/generate_mp_template.py

By default it writes to the mounted frontend public dir so the "MP Question
Template" download link resolves; pass --out to override.
"""

import argparse
from pathlib import Path

from openpyxl import Workbook

DEFAULT_OUT = "/app/frontend_public/templates/mp_template.xlsx"

QUESTION_HEADERS = [
    "misconception_tag",
    "text_en",
    "text_id",
    "code",
    "options_en",
    "options_id",
    "answer_index",
    "explanation_en",
    "explanation_id",
    # P5: which misconception(s) each distractor reveals — comma-separated tags
    # (KC family "VA" or specific code "VA-01"); leave the correct option blank.
    "option_a_misconceptions",
    "option_b_misconceptions",
    "option_c_misconceptions",
]

QUESTION_EXAMPLES = [
    [
        "VA",
        "After x <- 5 and then x <- x + 3, what is x?",
        "Setelah x <- 5 lalu x <- x + 3, berapa nilai x?",
        "x <- 5\nx <- x + 3",
        '["8", "3", "5"]',
        '["8", "3", "5"]',
        0,
        "x <- x + 3 reads the current value of x (5) and adds 3.",
        "x <- x + 3 membaca nilai x sekarang (5) lalu menambah 3.",
        "",              # option A is the correct answer — no triggers
        "VA-01",         # picking "3" reveals: assignment replaces instead of updates
        "VA-02, SQ",     # picking "5" reveals: second assignment ignored / order misread
    ],
    [
        "LO",
        "How many times does a while i < 3 loop with i starting at 0 and i <- i + 1 inside run?",
        "Berapa kali loop while i < 3 berjalan jika i mulai dari 0 dan ada i <- i + 1 di dalamnya?",
        "i <- 0\nwhile i < 3 do\n    i <- i + 1\nendwhile",
        '["3", "4", "2"]',
        '["3", "4", "2"]',
        0,
        "The loop body runs for i = 0, 1, 2 — three times; at i = 3 the condition fails.",
        "Badan loop berjalan untuk i = 0, 1, 2 — tiga kali; saat i = 3 kondisinya gagal.",
        "",
        "LO-01",         # off-by-one: thinks the failing check still executes the body
        "LO-02, CD",     # boundary misread: stops one early / condition misunderstood
    ],
]

MAPPING_HEADERS = ["problem_key", "misconception_tag"]
MAPPING_EXAMPLES = [
    ["simple-accumulator", "VA"],
    ["simple-accumulator", "LO-01"],
]


def main(out: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    ws = wb.create_sheet("question bank")
    ws.append(QUESTION_HEADERS)
    for row in QUESTION_EXAMPLES:
        ws.append(row)
    for i, h in enumerate(QUESTION_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(16, len(h) + 4)

    ws2 = wb.create_sheet("problem_misconceptions")
    ws2.append(MAPPING_HEADERS)
    for row in MAPPING_EXAMPLES:
        ws2.append(row)
    for i, h in enumerate(MAPPING_HEADERS, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = max(20, len(h) + 4)

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote MP question template -> {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=DEFAULT_OUT, help="Output .xlsx path")
    args = parser.parse_args()
    main(args.out)
