"""Generate a realistic mock Moodle quiz-report .xlsx for the LMS summary feature.

Mirrors the 6-sheet schema of a real export (COURSES, PARTICIPANTS, QUIZ_LIST,
QUESTION_BANK, PARTICIPANT_ATTEMPTS, RESPONSE_HISTORY) but with a brand-new course,
new students, and new quizzes so it imports cleanly alongside any existing data.

Unlike the real sample, QUESTION_BANK here includes the upcoming **Tag** column
(misconception codes like VA-01, LO-02) so the misconception panels populate.

Run inside the backend container:

    docker exec amt-backend python scripts/generate_mock_lms.py /tmp/mock.xlsx
"""

import random
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook

# ---- fixed identifiers (distinct from the real sample's course 18332) ----
COURSE_ID = 20001
COURSE_NAME = "ALGORITMA DAN PEMROGRAMAN 1 IF-50-INT [MOCK]"
SHORTNAME = "CAK1BAB3-IF-50-MOCK"
FAKULTAS = "FAKULTAS INFORMATIKA (FIF)"
PRODI = "PRODI S1 INFORMATIKA (FIF)"
TAHUN_AJAR = "2526/2"

# ---- problem bank: (tag, name, text, right_answer, wrong_answer) ----
P = {
    "VA-01": ("Variable Swapping", "Swap the values of two variables x and y.",
        "program swap\ndictionary\n    x, y, temp : integer\nalgorithm\n    read (x, y)\n    temp <- x\n    x <- y\n    y <- temp\n    write (x, y)\nendprogram",
        "program swap\ndictionary\n    x, y : integer\nalgorithm\n    read (x, y)\n    x <- y\n    y <- x\n    write (x, y)\nendprogram"),
    "CO-01": ("Circle Area", "Compute the area of a circle using constant PI.",
        "program circle\ndictionary\n    const PI = 3.14159\n    r, area : real\nalgorithm\n    read (r)\n    area <- PI * r * r\n    write (area)\nendprogram",
        "program circle\ndictionary\n    r, area : real\nalgorithm\n    read (r)\n    area <- 3 * r * r\n    write (area)\nendprogram"),
    "IO-01": ("Echo Message", "Read a message and print it back.",
        "program echo\ndictionary\n    msg : string\nalgorithm\n    read (msg)\n    write (msg)\nendprogram",
        "program echo\ndictionary\n    msg : string\nalgorithm\n    read (msg)\nendprogram"),
    "OP-01": ("Arithmetic Mix", "Compute y = (x + 3) * 2 - 5.",
        "program arith\ndictionary\n    x, y : integer\nalgorithm\n    read (x)\n    y <- (x + 3) * 2 - 5\n    write (y)\nendprogram",
        "program arith\ndictionary\n    x, y : integer\nalgorithm\n    read (x)\n    y <- x + 3 * 2 - 5\n    write (y)\nendprogram"),
    "LO-01": ("Sum 1..N", "Compute the sum 1 + 2 + ... + N.",
        "program sumn\ndictionary\n    n, i, total : integer\nalgorithm\n    read (n)\n    total <- 0\n    i <- 1\n    while i <= n do\n        total <- total + i\n        i <- i + 1\n    endwhile\n    write (total)\nendprogram",
        "program sumn\ndictionary\n    n, i, total : integer\nalgorithm\n    read (n)\n    total <- 0\n    i <- 1\n    if i <= n then\n        total <- total + i\n    endif\n    write (total)\nendprogram"),
    "LO-02": ("Count Down", "Print numbers from N down to 1.",
        "program countdown\ndictionary\n    n : integer\nalgorithm\n    read (n)\n    while n >= 1 do\n        write (n)\n        n <- n - 1\n    endwhile\nendprogram",
        "program countdown\ndictionary\n    n : integer\nalgorithm\n    read (n)\n    while n >= 1 do\n        write (n)\n        n <- n + 1\n    endwhile\nendprogram"),
    "EX-01": ("Average of Three", "Read three numbers and print their average.",
        "program avg3\ndictionary\n    a, b, c, avg : real\nalgorithm\n    read (a, b, c)\n    avg <- (a + b + c) / 3\n    write (avg)\nendprogram",
        "program avg3\ndictionary\n    a, b, c, avg : real\nalgorithm\n    read (a, b, c)\n    avg <- a + b + c / 3\n    write (avg)\nendprogram"),
    "CD-01": ("Even or Odd", "Print EVEN if n is even, else ODD.",
        "program evenodd\ndictionary\n    n : integer\nalgorithm\n    read (n)\n    if n mod 2 = 0 then\n        write (\"EVEN\")\n    else\n        write (\"ODD\")\n    endif\nendprogram",
        "program evenodd\ndictionary\n    n : integer\nalgorithm\n    read (n)\n    if n = 2 then\n        write (\"EVEN\")\n    else\n        write (\"ODD\")\n    endif\nendprogram"),
    "CD-02": ("Grade Classifier", "Print A if score >= 80, B if >= 60, else C.",
        "program grade\ndictionary\n    s : integer\nalgorithm\n    read (s)\n    if s >= 80 then\n        write (\"A\")\n    elseif s >= 60 then\n        write (\"B\")\n    else\n        write (\"C\")\n    endif\nendprogram",
        "program grade\ndictionary\n    s : integer\nalgorithm\n    read (s)\n    if s >= 80 then\n        write (\"A\")\n    else\n        write (\"C\")\n    endif\nendprogram"),
    "OP-02": ("Max of Two", "Print the larger of two numbers.",
        "program maxtwo\ndictionary\n    a, b : integer\nalgorithm\n    read (a, b)\n    if a > b then\n        write (a)\n    else\n        write (b)\n    endif\nendprogram",
        "program maxtwo\ndictionary\n    a, b : integer\nalgorithm\n    read (a, b)\n    if a < b then\n        write (a)\n    else\n        write (b)\n    endif\nendprogram"),
    "SQ-01": ("Celsius to Fahrenheit", "Convert Celsius to Fahrenheit: F = C * 9 / 5 + 32.",
        "program tempconv\ndictionary\n    c, f : real\nalgorithm\n    read (c)\n    f <- c * 9 / 5 + 32\n    write (f)\nendprogram",
        "program tempconv\ndictionary\n    c, f : real\nalgorithm\n    f <- c * 9 / 5 + 32\n    read (c)\n    write (f)\nendprogram"),
    "VA-02": ("Variable Assignment", "Read a and b, assign their sum to c, print c.",
        "program assign\ndictionary\n    a, b, c : integer\nalgorithm\n    read (a, b)\n    c <- a + b\n    write (c)\nendprogram",
        "program assign\ndictionary\n    a, b : integer\nalgorithm\n    read (a, b)\n    c <- a + b\n    write (c)\nendprogram"),
    "IO-02": ("Double Input Output", "Read two numbers and print both.",
        "program doubleio\ndictionary\n    a, b : integer\nalgorithm\n    read (a, b)\n    write (a, b)\nendprogram",
        "program doubleio\ndictionary\n    a, b : integer\nalgorithm\n    read (a)\n    write (a)\nendprogram"),
}

# ---- quizzes: (quiz_id, name, [tags], max_grade) ----
QUIZZES = [
    (30001, "Exercises 1A: Simple Data Types (Easy)", ["VA-01", "CO-01", "IO-01", "OP-01"], 10),
    (30002, "Exercises 2A: Iterations (Series)", ["LO-01", "LO-02", "EX-01"], 10),
    (30003, "Exercises 3A: Branching (Basic)", ["CD-01", "CD-02", "OP-02"], 10),
    (30004, "Assignment Proses Sekuensial", ["SQ-01", "VA-02", "IO-02"], 10),
]

# ---- students: (user_id, first, last, skill 0..1) ----
STUDENTS = [
    (50001, "ARYA", "PRATAMA", 0.90),
    (50002, "SITI", "NURHALIZA", 0.78),
    (50003, "BUDI", "SANTOSO", 0.50),
    (50004, "DEWI", "LESTARI", 0.82),
    (50005, "RIZKY", "FIRMANSYAH", 0.35),
    (50006, "PUTRI", "MAHARANI", 0.66),
    (50007, "FAJAR", "NUGROHO", 0.55),
    (50008, "AISYAH", "RAHMAWATI", 0.72),
]
TEACHER = (59001, "DOSEN", "PENGAMPU")

# The importer links an LMS participant to a local account when
# `users.username` == the participant's Username OR Email (see
# `_match_participants` in app/core/lms_import.py). Emitting the teacher's
# Username as the seeded demo instructor's local username auto-links this
# course to `instructor_user`, so the course-scoped (UC4) LMS Reports
# dashboard shows data immediately after import — no manual linking step.
# For a real export, set this to the teacher whose local username matches
# their Moodle username/email, or run scripts/seed_lms_teacher.py afterwards.
TEACHER_LOCAL_USERNAME = "instructor_user"

QUIZ_OPEN = datetime(2025, 9, 15, 8, 0, 0)
QUIZ_CLOSE = datetime(2025, 12, 20, 23, 59, 0)


def email(first, last):
    return f"{first.lower()}{last.lower().replace(' ', '')}@student.telkomuniversity.ac.id"


def teacher_email(first, last):
    return f"{first.lower()}{last.lower()}@telkomuniversity.ac.id"


def build(path):
    rng = random.Random(42)
    wb = Workbook()
    wb.remove(wb.active)

    ws_courses = wb.create_sheet("COURSES")
    ws_courses.append(["Fakultas", "Prodi", "Tahun_Ajar", "Course_ID", "Shortname", "Fullname"])
    ws_courses.append([FAKULTAS, PRODI, TAHUN_AJAR, COURSE_ID, SHORTNAME, COURSE_NAME])

    ws_part = wb.create_sheet("PARTICIPANTS")
    ws_part.append(["Course_ID", "Course_Name", "User_ID", "Username", "Firstname",
                    "Lastname", "Email", "Role_Name", "Role_Shortname"])
    tid, tf, tl = TEACHER
    # Username = local demo instructor so the importer matches this teacher to
    # `instructor_user`; Email stays the realistic Telkom address for display.
    ws_part.append([COURSE_ID, COURSE_NAME, tid, TEACHER_LOCAL_USERNAME, tf, tl,
                    teacher_email(tf, tl), "Teacher", "editingteacher"])
    for uid, first, last, _ in STUDENTS:
        e = email(first, last)
        ws_part.append([COURSE_ID, COURSE_NAME, uid, e, first, last, e, "Student", "student"])

    ws_quiz = wb.create_sheet("QUIZ_LIST")
    ws_quiz.append(["Course_ID", "Course_Name", "Quiz_ID", "Quiz_Name", "Open_Time",
                    "Close_Time", "Time_Limit_Seconds", "Max_Grade", "Total_Questions"])
    for qid, qname, tags, maxg in QUIZZES:
        ws_quiz.append([COURSE_ID, COURSE_NAME, qid, qname, QUIZ_OPEN, QUIZ_CLOSE,
                        1800, maxg, len(tags)])

    # QUESTION_BANK now carries a Tag column (the upcoming misconception tag).
    ws_qb = wb.create_sheet("QUESTION_BANK")
    ws_qb.append(["Course_ID", "Course_Name", "Quiz_ID", "Quiz_Name", "Slot_Number",
                  "Question_ID", "Question_Name", "Question_Type", "Question_Text", "Tag"])
    qid_of = {}
    for qi, (qid, qname, tags, _maxg) in enumerate(QUIZZES):
        for slot, tag in enumerate(tags, start=1):
            question_id = 40000 + qi * 100 + slot
            qid_of[(qid, slot)] = question_id
            name, text, _r, _w = P[tag]
            ws_qb.append([COURSE_ID, COURSE_NAME, qid, qname, slot, question_id,
                          name, "coderunner", text, tag])

    ws_pa = wb.create_sheet("PARTICIPANT_ATTEMPTS")
    ws_pa.append(["Course_ID", "Course_Name", "Quiz_ID", "Quiz_Name", "User_ID", "Username",
                  "Firstname", "Lastname", "Email", "Attempt_Number", "Attempt_Start_Time",
                  "Attempt_Finish_Time", "Attempt_Total_Grade", "Slot_Number", "Question_Summary",
                  "Right_Answer", "Student_Answer", "Question_State", "Question_Grade"])

    ws_rh = wb.create_sheet("RESPONSE_HISTORY")
    ws_rh.append(["Course_ID", "Course_Name", "Quiz_ID", "Quiz_Name", "User_ID", "Username",
                  "Firstname", "Lastname", "Email", "Attempt_Number", "Slot_Number",
                  "Question_ID", "Step", "Time", "Action", "State", "Marks", "CodeRunner_Output"])

    for uid, first, last, skill in STUDENTS:
        e = email(first, last)
        for qid, qname, tags, maxg in QUIZZES:
            n_attempts = 1 + (1 if skill < 0.6 else 0) + (1 if rng.random() < 0.35 else 0)
            start_day = QUIZ_OPEN + timedelta(days=rng.randint(0, 80))
            for attempt in range(1, n_attempts + 1):
                a_start = start_day + timedelta(days=(attempt - 1) * rng.randint(1, 7),
                                                hours=rng.randint(0, 10), minutes=rng.randint(0, 59))
                correct_count = 0
                slot_results = []
                for slot, tag in enumerate(tags, start=1):
                    name, text, right, wrong = P[tag]
                    p_correct = min(0.97, skill + 0.12 * (attempt - 1))
                    is_correct = rng.random() < p_correct
                    if is_correct:
                        state, grade, answer = "gradedright", 1.0, right
                        correct_count += 1
                    elif rng.random() < 0.12:
                        state, grade, answer = "gaveup", 0.0, wrong
                    else:
                        state, grade, answer = "gradedwrong", 0.0, wrong
                    slot_results.append((slot, tag, name, text, right, answer, state, grade))

                total_grade = round(correct_count / len(tags) * maxg, 2)
                a_finish = a_start + timedelta(minutes=rng.randint(6, 40))

                for slot, tag, name, text, right, answer, state, grade in slot_results:
                    ws_pa.append([COURSE_ID, COURSE_NAME, qid, qname, uid, e, first, last, e,
                                  attempt, a_start, a_finish, total_grade, slot,
                                  text, right, answer, state, grade])

                    question_id = qid_of[(qid, slot)]
                    first_line = answer.split("\n")[0]
                    t0 = a_start + timedelta(minutes=slot * 2)
                    t1 = t0 + timedelta(minutes=rng.randint(1, 4))
                    ws_rh.append([COURSE_ID, COURSE_NAME, qid, qname, uid, e, first, last, e,
                                  attempt, slot, question_id, 1, a_start, "Started",
                                  "Not complete", None, None])
                    ws_rh.append([COURSE_ID, COURSE_NAME, qid, qname, uid, e, first, last, e,
                                  attempt, slot, question_id, 2, t0, f"Submit: {first_line}",
                                  "Complete", grade,
                                  "All tests passed" if grade > 0 else "2 of 3 tests failed"])
                    ws_rh.append([COURSE_ID, COURSE_NAME, qid, qname, uid, e, first, last, e,
                                  attempt, slot, question_id, 3, t1,
                                  "Attempt finished submitting: {$a}",
                                  "Correct" if grade > 0 else "Incorrect", grade,
                                  "All tests passed" if grade > 0 else "2 of 3 tests failed"])

    wb.save(path)
    print(f"Wrote {path}")
    print(f"  course {COURSE_ID}: {len(STUDENTS)} students, {len(QUIZZES)} quizzes, "
          f"{sum(len(t) for _, _, t, _ in QUIZZES)} tagged questions")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "/tmp/Moodle_Quiz_Report_MOCK.xlsx"
    build(out)
