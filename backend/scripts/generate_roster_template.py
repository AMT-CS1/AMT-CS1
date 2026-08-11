"""Generate the researcher roster-import template (R4).

Writes an .xlsx with COURSES / TEACHERS / STUDENTS sheets and one example row
each, matching what `app/core/roster_import.py` parses.

Run inside the backend container (openpyxl lives there):

    docker exec amt-backend python scripts/generate_roster_template.py

By default it writes to the mounted frontend public dir so the researcher page's
"Template" link resolves; pass --out to override.
"""

import argparse
from pathlib import Path

from openpyxl import Workbook

DEFAULT_OUT = "/app/frontend_public/templates/roster_template.xlsx"

SHEETS = {
    "COURSES": (
        ["Course_ID", "Shortname", "Fullname", "Fakultas", "Prodi", "Tahun_Ajar"],
        [101, "CS1-2026", "Introduction to Programming", "Ilmu Komputer", "Informatika", "2025/2026"],
    ),
    "TEACHERS": (
        ["Course_ID", "Username", "Full_Name", "Email", "LMS_User_ID", "Password"],
        [101, "dosen_budi", "Budi Santoso", "budi@example.com", 5001, ""],
    ),
    "STUDENTS": (
        ["Course_ID", "Username", "Full_Name", "Email", "LMS_User_ID", "Password"],
        [101, "student_ani", "Ani Wijaya", "ani@example.com", 9001, ""],
    ),
}


def main(out: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    for name, (headers, example) in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        ws.append(example)
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(h) + 4)

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote roster template -> {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=DEFAULT_OUT, help="Output .xlsx path")
    args = parser.parse_args()
    main(args.out)
